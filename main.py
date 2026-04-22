import asyncio, csv, io, json, os, shutil, uuid
import httpx
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query, BackgroundTasks
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import config
from pipeline.loader import load_folder, load_image, SUPPORTED
from pipeline.extractor import extract_batch
import database as db
import vectorstore as vs
import chat
import memory
from pydantic import BaseModel

app = FastAPI(title="KONTACT — Catalog Vision RAG")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

FRONTEND_BUILD = os.path.join(os.path.dirname(__file__), "frontend", "build")


from typing import Optional

class ChatRequest(BaseModel):
    question: str
    session_id: Optional[str] = None


# ─── UPLOAD (just saves to queue, no processing) ───

@app.post("/api/upload")
async def upload_images(files: list[UploadFile] = File(...), batch_id: str = Form(None), bg: BackgroundTasks = BackgroundTasks()):
    if not batch_id:
        batch_id = str(uuid.uuid4())[:8]
    batch_dir = os.path.join(config.UPLOADS_DIR, batch_id)
    os.makedirs(batch_dir, exist_ok=True)

    queued = []
    for f in files:
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in SUPPORTED:
            continue
        dest = os.path.join(batch_dir, f.filename)
        with open(dest, "wb") as out:
            shutil.copyfileobj(f.file, out)
        if ext == ".pdf":
            # Convert PDF pages to individual JPEG images
            import fitz
            from PIL import Image as PILImage
            doc = fitz.open(dest)
            base_name = os.path.splitext(f.filename)[0]
            for i, page in enumerate(doc, start=1):
                pix = page.get_pixmap(dpi=200)
                img = PILImage.frombytes("RGB", [pix.width, pix.height], pix.samples)
                page_name = f"{base_name}_page{i}.jpg"
                page_dest = os.path.join(batch_dir, page_name)
                img.save(page_dest, format="JPEG", quality=90)
                db.queue_add(batch_id, page_name, page_dest)
                queued.append(page_name)
            doc.close()
        else:
            db.queue_add(batch_id, f.filename, dest)
            queued.append(f.filename)

    # Auto-process in background after upload
    if queued:
        bg.add_task(_process_batch, batch_id)

    return {"batch_id": batch_id, "queued": len(queued), "files": queued, "auto_processing": True}


@app.post("/api/upload/folder")
def upload_from_folder(folder_path: str):
    if not os.path.isdir(folder_path):
        raise HTTPException(400, f"Not a folder: {folder_path}")
    batch_id = os.path.basename(folder_path.rstrip("/"))
    count = 0
    for f in sorted(os.listdir(folder_path)):
        ext = os.path.splitext(f)[1].lower()
        if ext in SUPPORTED:
            db.queue_add(batch_id, f, os.path.join(folder_path, f))
            count += 1
    if not count:
        raise HTTPException(400, "No supported images found")
    return {"batch_id": batch_id, "queued": count}


# ─── QUEUE STATUS ───

@app.get("/api/queue")
def queue_status(batch_id: str = None):
    if batch_id:
        return db.queue_status(batch_id)
    return db.queue_status()


@app.get("/api/queue/batches")
def queue_batches():
    return db.queue_batches()


@app.get("/api/queue/errors")
def queue_errors(batch_id: str = None):
    return db.queue_errors(batch_id)


@app.post("/api/queue/retry/{queue_id}")
def retry_queue_item(queue_id: int):
    success = db.queue_retry(queue_id)
    if not success:
        raise HTTPException(404, "Item not found or not in error state")
    return {"retried": queue_id}


@app.delete("/api/batch/{batch_id}")
def delete_batch(batch_id: str):
    count = db.delete_batch(batch_id)
    if count == 0:
        raise HTTPException(404, "Batch not found")
    return {"deleted": batch_id, "count": count}


@app.get("/api/queue/pending")
def queue_pending(batch_id: str = None):
    return db.queue_pending(batch_id)


# ─── PROCESS (runs extraction on pending queue items) ───

async def _process_batch(batch_id: str = None):
    pending = db.queue_pending(batch_id)
    if not pending:
        return {"processed": 0}

    images = []
    queue_map = {}
    for item in pending:
        try:
            img = load_image(item["file_path"])
            images.append(img)
            queue_map[img["file"]] = item
        except Exception as e:
            db.queue_update(item["id"], "error", error=str(e))

    if not images:
        return {"processed": 0, "errors": len(pending)}

    def on_progress(done, total, name):
        print(f"  [{done}/{total}] {name}")

    results = await extract_batch(images, on_progress=on_progress)

    done_count = 0
    for r in results:
        fname = r.get("source_file", "")
        item = queue_map.get(fname)
        if not item:
            continue
        if r.get("error"):
            db.queue_update(item["id"], "error", error=r["error"])
        else:
            db.queue_update(item["id"], "done", image_type=r.get("image_type"))
            folder = item["batch_id"]
            db.insert_extraction(folder, r)
            vs.index_record(folder, r)
            done_count += 1

    # Save JSON
    bid = batch_id or "batch"
    out_path = os.path.join(config.EXTRACTIONS_DIR, f"{bid}.json")
    with open(out_path, "w") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    return {"processed": done_count, "errors": len(results) - done_count}


@app.post("/api/process")
async def process_queue(batch_id: str = None):
    result = await _process_batch(batch_id)
    return result


@app.post("/api/process/background")
async def process_background(batch_id: str = None, bg: BackgroundTasks = None):
    pending = db.queue_pending(batch_id)
    if not pending:
        return {"status": "complete", "pending": 0, "batch_id": batch_id, "message": "Nothing pending"}
    bg.add_task(_process_batch, batch_id)
    return {"status": "processing", "pending": len(pending), "batch_id": batch_id}


# ─── SEARCH ───

@app.get("/api/search")
def search(q: str = Query(...), limit: int = 20):
    return db.search(q, limit)


@app.get("/api/search/semantic")
def semantic_search(q: str = Query(...), n: int = 5):
    return vs.query(q, n)


# ─── DATA ───

@app.get("/api/data")
def get_data(folder: str = None, limit: int = 500):
    if folder:
        return db.get_by_folder(folder)
    return db.get_all(limit)


@app.get("/api/stats")
def stats():
    return db.get_stats()


@app.post("/api/index")
def index_all():
    db_count = db.load_all_extractions()
    vs_count = vs.index_all_from_json()
    return {"db_records": db_count, "vector_records": vs_count}


# ─── CHAT ───

@app.post("/api/chat")
async def chat_endpoint(req: ChatRequest):
    session_id = req.session_id or str(uuid.uuid4())[:8]
    result = await chat.ask(req.question, session_id=session_id)
    return result


@app.get("/api/chat/sessions")
def chat_sessions():
    return db.list_sessions()


@app.get("/api/chat/history/{session_id}")
def chat_history(session_id: str):
    return db.get_chat_history(session_id)


@app.delete("/api/chat/sessions/{session_id}")
def delete_chat_session(session_id: str):
    db.delete_session(session_id)
    return {"deleted": session_id}


# ─── EXPORT ───

@app.get("/api/export/json")
def export_json():
    data = db.export_all()
    return JSONResponse(content=data, headers={"Content-Disposition": "attachment; filename=kontact_export.json"})


@app.get("/api/contacts")
def get_contacts():
    return db.get_contacts_table()


@app.get("/api/products")
def get_products():
    return db.get_products_table()


@app.get("/api/documents/metadata")
def get_documents_metadata(limit: int = 500):
    return db.get_documents_with_metadata(limit)


@app.post("/api/migrate")
def run_migration():
    result = db.populate_normalized_tables()
    return {"status": "ok", **result}


@app.get("/api/dashboard")
def dashboard():
    return db.get_dashboard_stats()


@app.get("/api/export/xlsx")
def export_xlsx():
    from openpyxl import Workbook
    data = db.export_all()
    if not data:
        raise HTTPException(404, "No data")

    wb = Workbook()

    # Sheet 1: Products
    ws_prod = wb.active
    ws_prod.title = "Products"
    prod_headers = ["company", "product_name", "model", "specs", "category", "price", "folder", "source_file"]
    ws_prod.append(prod_headers)
    for row in data:
        prods = row.get("products", [])
        if isinstance(prods, str):
            try:
                prods = json.loads(prods)
            except Exception:
                prods = []
        if not isinstance(prods, list):
            continue
        for p in prods:
            if not isinstance(p, dict):
                continue
            ws_prod.append([
                p.get("company", "") or row.get("company", ""),
                p.get("product_name", "") or p.get("name", ""),
                p.get("model", ""),
                p.get("specs", "") if isinstance(p.get("specs"), str) else json.dumps(p.get("specs", ""), ensure_ascii=False),
                p.get("category", ""),
                p.get("price", ""),
                row.get("folder", ""),
                row.get("source_file", ""),
            ])

    # Sheet 2: Contacts
    ws_cont = wb.create_sheet("Contacts")
    cont_headers = ["company", "person", "phone", "email", "website", "address", "folder"]
    ws_cont.append(cont_headers)
    contacts = db.get_all_contacts()
    for ct in contacts:
        ws_cont.append([ct.get(h, "") for h in cont_headers])

    # Sheet 3: Summary
    ws_sum = wb.create_sheet("Summary")
    sum_headers = ["company", "document_count", "product_count", "has_contact"]
    ws_sum.append(sum_headers)
    company_info = {}
    contact_companies = {ct["company"] for ct in contacts if ct.get("company")}
    for row in data:
        comp = row.get("company", "") or "Unknown"
        if comp not in company_info:
            company_info[comp] = {"doc_count": 0, "prod_count": 0}
        company_info[comp]["doc_count"] += 1
        prods = row.get("products", [])
        if isinstance(prods, str):
            try:
                prods = json.loads(prods)
            except Exception:
                prods = []
        if isinstance(prods, list):
            company_info[comp]["prod_count"] += len(prods)
    for comp, info in sorted(company_info.items()):
        ws_sum.append([comp, info["doc_count"], info["prod_count"], "Yes" if comp in contact_companies else "No"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kontact_export.xlsx"},
    )


@app.get("/api/export/csv")
def export_csv():
    data = db.export_all()
    if not data:
        raise HTTPException(404, "No data")
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["folder", "source_file", "image_type", "company", "title", "raw_text"])
    writer.writeheader()
    for row in data:
        writer.writerow({k: row.get(k, "") for k in writer.fieldnames})
    output.seek(0)
    return StreamingResponse(output, media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=kontact_export.csv"})


# ─── IMAGE SERVING (for citations) ───

@app.get("/api/image/{folder}/{filename:path}")
async def serve_image(folder: str, filename: str):
    from fastapi.responses import FileResponse
    if '..' in folder or '..' in filename or folder.startswith('/') or filename.startswith('/'):
        raise HTTPException(400, "Invalid path")
    try:
        # Try uploads dir
        path = os.path.join(config.UPLOADS_DIR, folder, filename)
        if os.path.isfile(path):
            return FileResponse(path)
        # Try queue file_path
        c = db._conn()
        row = c.execute("SELECT file_path FROM queue WHERE batch_id = ? AND file_name = ?", (folder, filename)).fetchone()
        if row:
            fp = row["file_path"]
            c.close()
            if fp and os.path.isfile(fp):
                return FileResponse(fp)
        else:
            row = c.execute("SELECT source_path FROM documents WHERE folder = ? AND source_file = ?", (folder, filename)).fetchone()
            c.close()
            if row:
                fp = row["source_path"]
                if fp and os.path.isfile(fp):
                    return FileResponse(fp)
    except Exception as e:
        print(f"Image serve error: {e} for {folder}/{filename}")
    raise HTTPException(404, "Image not found")


# ─── STREAMING CHAT ───

@app.post("/api/chat/stream")
async def chat_stream(req: ChatRequest):
    from starlette.responses import StreamingResponse as SSEResponse
    from tools import execute_tool
    session_id = req.session_id or str(uuid.uuid4())[:8]

    async def event_generator():
        # Send session info
        yield f"event: session\ndata: {json.dumps({'session_id': session_id})}\n\n"

        # Build context with learning memory
        yield f"event: status\ndata: {json.dumps({'step': 'Searching knowledge base...'})}\n\n"
        context = chat._build_context(req.question)
        system_prompt = chat._build_system_prompt()

        yield f"event: status\ndata: {json.dumps({'step': 'Querying LLM...'})}\n\n"

        # Get history
        history = db.get_chat_history(session_id, limit=6) if session_id else []

        messages = [{"role": "system", "content": system_prompt}]
        for h in history[-6:]:
            messages.append({"role": h["role"], "content": h["content"]})
        messages.append({"role": "user", "content": f"Context:\n{context}\n\nQuestion: {req.question}"})

        _stream_headers = {
            "Authorization": f"Bearer {config.OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
        }

        async def _do_stream(msgs):
            """Async generator: stream LLM call, yielding content deltas."""
            payload = {
                "model": config.VISION_MODEL,
                "messages": msgs,
                "max_tokens": 2000,
                "temperature": 0.1,
                "stream": True,
            }
            async with httpx.AsyncClient() as client:
                async with client.stream("POST", config.OPENROUTER_BASE,
                                         json=payload, headers=_stream_headers, timeout=120) as r:
                    async for line in r.aiter_lines():
                        if not line.startswith("data: "):
                            continue
                        data_str = line[6:]
                        if data_str.strip() == "[DONE]":
                            break
                        try:
                            chunk = json.loads(data_str)
                            delta = chunk.get("choices", [{}])[0].get("delta", {}).get("content", "")
                            if delta:
                                yield delta
                        except Exception:
                            continue

        # First LLM call — buffer it (don't stream yet, might contain tool calls)
        full_answer = ""
        async for delta in _do_stream(messages):
            full_answer += delta

        # Tool execution loop (up to MAX_TOOL_ITERATIONS)
        for _tool_iter in range(chat.MAX_TOOL_ITERATIONS):
            if not chat._has_tool_calls(full_answer):
                break

            import time as _time
            tool_calls = chat._parse_tool_calls(full_answer)
            tool_results = []
            for tc in tool_calls:
                tname = tc["name"]
                yield f"event: status\ndata: {json.dumps({'step': 'Running ' + tname + '...'})}\n\n"
                t0 = _time.time()
                r = execute_tool(tname, tc["args"])
                dur = round(_time.time() - t0, 2)
                step_done = '✓ ' + tname + ' (' + str(dur) + 's)'
                yield f"event: status\ndata: {json.dumps({'step': step_done})}\n\n"
                tool_results.append(f"[RESULT: {tname}]\n{r}\n[/RESULT]")

            result_block = "\n\n".join(tool_results)
            messages.append({"role": "assistant", "content": full_answer})
            messages.append({"role": "user", "content": (
                f"Tool results:\n\n{result_block}\n\n"
                "Now provide your final answer. Do NOT include [TOOL:] markers. "
                "Just give the clean answer with data from the tool results."
            )})

            yield f"event: status\ndata: {json.dumps({'step': 'Generating answer...'})}\n\n"

            # Get final answer — buffer again in case of more tool calls
            full_answer = ""
            async for delta in _do_stream(messages):
                full_answer += delta

        # Strip any leftover tool markers
        full_answer = chat._TOOL_PATTERN.sub("", full_answer).strip()

        # Now stream the clean final answer to the client
        yield f"event: content\ndata: {json.dumps({'text': full_answer})}\n\n"

        # Get sources
        sources = []
        for s in vs.query(req.question, n_results=3):
            src_file = s["metadata"]["source_file"]
            folder = s["metadata"]["folder"]
            sources.append({
                "folder": folder,
                "file": src_file,
                "company": s["metadata"].get("company", ""),
                "image_url": f"/api/image/{folder}/{src_file}",
            })

        # Save to history
        db.save_chat(session_id, "user", req.question)
        db.save_chat(session_id, "assistant", full_answer)

        yield f"event: done\ndata: {json.dumps({'sources': sources, 'session_id': session_id})}\n\n"

    return SSEResponse(event_generator(), media_type="text/event-stream")


# ─── FEEDBACK & MEMORY ───

class FeedbackRequest(BaseModel):
    session_id: str
    question: str
    answer: str
    rating: str  # "up" or "down"


@app.post("/api/feedback")
def submit_feedback(req: FeedbackRequest):
    memory.save_feedback(req.session_id, req.question, req.answer, req.rating)
    return {"saved": True, "rating": req.rating}


@app.get("/api/memories")
def list_memories(limit: int = 10):
    return memory.get_memories(limit=limit)


@app.get("/api/config")
def get_config():
    return {
        "vision_model": config.VISION_MODEL,
        "embedding_model": config.EMBEDDING_MODEL,
        "provider": "OpenRouter",
    }


@app.get("/health")
def health():
    return {"status": "ok"}


# Serve frontend — static assets + SPA fallback
if os.path.isdir(FRONTEND_BUILD):
    from fastapi.responses import FileResponse
    app.mount("/_app", StaticFiles(directory=os.path.join(FRONTEND_BUILD, "_app")), name="assets")

    @app.get("/{path:path}")
    async def spa_fallback(path: str):
        file_path = os.path.join(FRONTEND_BUILD, path)
        if os.path.isfile(file_path):
            return FileResponse(file_path)
        return FileResponse(os.path.join(FRONTEND_BUILD, "index.html"))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
